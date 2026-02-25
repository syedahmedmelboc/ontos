"""
Bulk Asset Import/Export Manager

Handles CSV/XLSX export of assets and bulk import with preview,
validation, cycle detection, duplicate detection, and size caps.
"""

import csv
import io
import json
import os
from collections import defaultdict
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple
from uuid import UUID

from openpyxl import Workbook
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from src.common.logging import get_logger
from src.controller.assets_manager import AssetsManager
from src.db_models.assets import AssetDb, AssetRelationshipDb
from src.models.assets import (
    AssetCreate,
    AssetStatus,
    AssetUpdate,
)
from src.repositories.assets_repository import (
    asset_repo,
    asset_relationship_repo,
    asset_type_repo,
)

logger = get_logger(__name__)

BULK_IMPORT_MAX_ROWS = int(os.environ.get("BULK_IMPORT_MAX_ROWS", "10000"))
MAX_RELATIONSHIP_DEPTH = 10

EXPORT_COLUMNS = [
    "id",
    "name",
    "asset_type",
    "description",
    "platform",
    "location",
    "domain_id",
    "status",
    "tags",
    "properties",
    "parent_asset",
    "parent_relationship_type",
    "created_by",
    "created_at",
    "updated_at",
]

IMPORT_COLUMNS = [
    "id",
    "name",
    "asset_type",
    "description",
    "platform",
    "location",
    "domain_id",
    "status",
    "tags",
    "properties",
    "parent_asset",
    "parent_relationship_type",
]


class ImportAction(str, Enum):
    CREATE = "create"
    UPDATE = "update"
    SKIP = "skip"
    ERROR = "error"


class ImportPreviewItem(BaseModel):
    row: int = Field(..., description="1-based row number in the file")
    name: str = Field("", description="Asset name from the row")
    asset_type: str = Field("", description="Asset type name from the row")
    action: ImportAction = Field(..., description="What will happen with this row")
    message: Optional[str] = Field(None, description="Error or info message")
    existing_asset_id: Optional[str] = Field(None, description="ID of existing asset if updating")


class ImportPreviewResult(BaseModel):
    total_rows: int = 0
    will_create: int = 0
    will_update: int = 0
    will_skip: int = 0
    errors: int = 0
    items: List[ImportPreviewItem] = Field(default_factory=list)
    error_messages: List[str] = Field(default_factory=list)


class ImportResultItem(BaseModel):
    row: int
    name: str = ""
    asset_type: str = ""
    action: ImportAction
    asset_id: Optional[str] = None
    message: Optional[str] = None


class ImportResult(BaseModel):
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    items: List[ImportResultItem] = Field(default_factory=list)
    error_messages: List[str] = Field(default_factory=list)


def _parse_tags(raw: Optional[str]) -> Optional[List[str]]:
    if not raw or not raw.strip():
        return None
    return [t.strip() for t in raw.split(";") if t.strip()]


def _format_tags(tags: Optional[list]) -> str:
    if not tags:
        return ""
    return ";".join(str(t) for t in tags)


def _parse_properties(raw: Optional[str]) -> Optional[Dict[str, Any]]:
    if not raw or not raw.strip():
        return None
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        raise ValueError(f"Invalid JSON in properties: {raw[:100]}")


def _format_properties(props: Optional[dict]) -> str:
    if not props:
        return ""
    return json.dumps(props, ensure_ascii=False)


class AssetBulkManager:
    def __init__(self, assets_manager: Optional[AssetsManager] = None):
        self._assets_manager = assets_manager
        self._type_repo = asset_type_repo
        self._asset_repo = asset_repo
        self._rel_repo = asset_relationship_repo
        logger.debug("AssetBulkManager initialized.")

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_assets(
        self,
        db: Session,
        *,
        fmt: str = "csv",
        asset_ids: Optional[List[UUID]] = None,
        asset_type_id: Optional[UUID] = None,
        platform: Optional[str] = None,
        domain_id: Optional[str] = None,
        status: Optional[str] = None,
    ) -> Tuple[bytes, str, str]:
        """Export filtered assets. Returns (file_bytes, filename, content_type).

        When asset_ids is provided, only those specific assets are exported
        and other filters are ignored.
        """
        from sqlalchemy.orm import selectinload

        query = db.query(AssetDb).options(selectinload(AssetDb.asset_type)).order_by(AssetDb.name)

        if asset_ids:
            query = query.filter(AssetDb.id.in_(asset_ids))
        else:
            if asset_type_id:
                query = query.filter(AssetDb.asset_type_id == asset_type_id)
            if platform:
                query = query.filter(AssetDb.platform == platform)
            if domain_id:
                query = query.filter(AssetDb.domain_id == domain_id)
            if status:
                query = query.filter(AssetDb.status == status)

        assets = query.limit(BULK_IMPORT_MAX_ROWS).all()
        rows = self._assets_to_rows(assets)

        if fmt == "xlsx":
            return self._rows_to_xlsx(rows), "assets-export.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return self._rows_to_csv(rows), "assets-export.csv", "text/csv"

    def export_template(
        self,
        db: Session,
        *,
        asset_type_name: Optional[str] = None,
        fmt: str = "csv",
    ) -> Tuple[bytes, str, str]:
        """Export an empty template with headers and an example row."""
        example = {col: "" for col in IMPORT_COLUMNS}
        example["name"] = "Example Asset"
        example["status"] = "draft"

        if asset_type_name:
            example["asset_type"] = asset_type_name
            db_type = self._type_repo.get_by_name(db, name=asset_type_name)
            if db_type and db_type.required_fields:
                props = {}
                for field_name, field_def in db_type.required_fields.items():
                    if isinstance(field_def, dict):
                        props[field_name] = f"<{field_def.get('type', 'value')}>"
                    else:
                        props[field_name] = "<value>"
                example["properties"] = json.dumps(props, ensure_ascii=False)
        else:
            example["asset_type"] = "Table"

        rows = [example]
        if fmt == "xlsx":
            return self._rows_to_xlsx(rows, columns=IMPORT_COLUMNS), "assets-template.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return self._rows_to_csv(rows, columns=IMPORT_COLUMNS), "assets-template.csv", "text/csv"

    def _assets_to_rows(self, assets: List[AssetDb]) -> List[Dict[str, str]]:
        rows = []
        for a in assets:
            rows.append({
                "id": str(a.id),
                "name": a.name or "",
                "asset_type": a.asset_type.name if a.asset_type else "",
                "description": a.description or "",
                "platform": a.platform or "",
                "location": a.location or "",
                "domain_id": a.domain_id or "",
                "status": a.status or "",
                "tags": _format_tags(a.tags),
                "properties": _format_properties(a.properties),
                "parent_asset": "",
                "parent_relationship_type": "",
                "created_by": a.created_by or "",
                "created_at": str(a.created_at) if a.created_at else "",
                "updated_at": str(a.updated_at) if a.updated_at else "",
            })
        return rows

    def _rows_to_csv(self, rows: List[Dict[str, str]], columns: Optional[List[str]] = None) -> bytes:
        cols = columns or EXPORT_COLUMNS
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({c: row.get(c, "") for c in cols})
        return buf.getvalue().encode("utf-8")

    def _rows_to_xlsx(self, rows: List[Dict[str, str]], columns: Optional[List[str]] = None) -> bytes:
        cols = columns or EXPORT_COLUMNS
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets"
        ws.append(cols)
        for row in rows:
            ws.append([row.get(c, "") for c in cols])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Import - Parse
    # ------------------------------------------------------------------

    def _parse_file(self, file_bytes: bytes, filename: str) -> List[Dict[str, str]]:
        """Parse CSV or XLSX into list of row dicts."""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext in ("xlsx", "xls"):
            return self._parse_xlsx(file_bytes)
        return self._parse_csv(file_bytes)

    def _parse_csv(self, data: bytes) -> List[Dict[str, str]]:
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            cleaned = {k.strip().lower(): (v or "").strip() for k, v in row.items() if k}
            rows.append(cleaned)
        return rows

    def _parse_xlsx(self, data: bytes) -> List[Dict[str, str]]:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip().lower() if c else "" for c in next(rows_iter)]
        except StopIteration:
            wb.close()
            return []

        rows = []
        for row_values in rows_iter:
            row = {}
            for i, val in enumerate(row_values):
                if i < len(header) and header[i]:
                    row[header[i]] = str(val).strip() if val is not None else ""
            rows.append(row)

        wb.close()
        return rows

    # ------------------------------------------------------------------
    # Import - Validate & Preview
    # ------------------------------------------------------------------

    def _resolve_asset_types(self, db: Session) -> Dict[str, UUID]:
        """Build name->id map for all asset types."""
        all_types = self._type_repo.get_multi_filtered(db, limit=10000)
        return {t.name.lower(): t.id for t in all_types}

    def _detect_duplicates(self, rows: List[Dict[str, str]]) -> Dict[str, int]:
        """Find duplicate identity tuples within the file. Returns identity_key -> first_row (1-based)."""
        seen: Dict[str, int] = {}
        duplicates: Dict[str, int] = {}
        for idx, row in enumerate(rows):
            key = (
                row.get("name", "").lower(),
                row.get("asset_type", "").lower(),
                row.get("platform", "").lower(),
                row.get("location", "").lower(),
            )
            key_str = "|".join(key)
            if key_str in seen:
                duplicates[key_str] = seen[key_str]
            else:
                seen[key_str] = idx + 1
        return duplicates

    def _detect_parent_cycles(self, rows: List[Dict[str, str]]) -> List[str]:
        """Check for cycles in parent_asset references within the file."""
        graph: Dict[str, List[str]] = defaultdict(list)
        nodes = set()

        for row in rows:
            name = row.get("name", "").strip()
            parent = row.get("parent_asset", "").strip()
            if name and parent:
                graph[parent].append(name)
                nodes.add(name)
                nodes.add(parent)

        # Topological sort via DFS to detect cycles
        WHITE, GRAY, BLACK = 0, 1, 2
        color = {n: WHITE for n in nodes}
        cycles = []

        def dfs(node: str, path: List[str]) -> bool:
            color[node] = GRAY
            for child in graph.get(node, []):
                if color.get(child, WHITE) == GRAY:
                    cycle_start = path.index(child) if child in path else 0
                    cycle_path = path[cycle_start:] + [child]
                    cycles.append(" -> ".join(cycle_path))
                    return True
                if color.get(child, WHITE) == WHITE:
                    if dfs(child, path + [child]):
                        return True
            color[node] = BLACK
            return False

        for node in nodes:
            if color.get(node, WHITE) == WHITE:
                dfs(node, [node])

        return cycles

    def preview_import(
        self, db: Session, *, file_bytes: bytes, filename: str
    ) -> ImportPreviewResult:
        """Parse and validate file, return preview of what would happen."""
        result = ImportPreviewResult()

        rows = self._parse_file(file_bytes, filename)
        result.total_rows = len(rows)

        if len(rows) == 0:
            result.error_messages.append("File is empty or has no data rows.")
            return result

        if len(rows) > BULK_IMPORT_MAX_ROWS:
            result.error_messages.append(
                f"File contains {len(rows)} rows, maximum is {BULK_IMPORT_MAX_ROWS}. Split into multiple files."
            )
            return result

        type_map = self._resolve_asset_types(db)

        # Duplicate detection within file
        dup_first_rows = self._detect_duplicates(rows)
        dup_keys = set()
        for idx, row in enumerate(rows):
            key = "|".join([
                row.get("name", "").lower(),
                row.get("asset_type", "").lower(),
                row.get("platform", "").lower(),
                row.get("location", "").lower(),
            ])
            if key in dup_first_rows and (idx + 1) != dup_first_rows[key]:
                dup_keys.add((idx, key))

        # Cycle detection
        cycles = self._detect_parent_cycles(rows)
        if cycles:
            result.error_messages.append(
                f"Circular parent references detected: {'; '.join(cycles)}"
            )

        for idx, row in enumerate(rows):
            row_num = idx + 1
            name = row.get("name", "").strip()
            asset_type_raw = row.get("asset_type", "").strip()
            item = ImportPreviewItem(row=row_num, name=name, asset_type=asset_type_raw, action=ImportAction.ERROR)

            if not name:
                item.message = "Missing required field: name"
                result.errors += 1
                result.items.append(item)
                continue

            if not asset_type_raw:
                item.message = "Missing required field: asset_type"
                result.errors += 1
                result.items.append(item)
                continue

            type_id = type_map.get(asset_type_raw.lower())
            if not type_id:
                item.message = f"Unknown asset type: '{asset_type_raw}'"
                result.errors += 1
                result.items.append(item)
                continue

            # Check for duplicate in file
            dup_key = "|".join([
                name.lower(), asset_type_raw.lower(),
                row.get("platform", "").strip().lower(),
                row.get("location", "").strip().lower(),
            ])
            if (idx, dup_key) in dup_keys:
                item.message = f"Duplicate of row {dup_first_rows[dup_key]} in this file"
                item.action = ImportAction.ERROR
                result.errors += 1
                result.items.append(item)
                continue

            # Validate status
            status_raw = row.get("status", "").strip().lower()
            if status_raw and status_raw not in [s.value for s in AssetStatus]:
                item.message = f"Invalid status: '{status_raw}'. Must be one of: {', '.join(s.value for s in AssetStatus)}"
                result.errors += 1
                result.items.append(item)
                continue

            # Validate properties JSON
            try:
                _parse_properties(row.get("properties", ""))
            except ValueError as e:
                item.message = str(e)
                result.errors += 1
                result.items.append(item)
                continue

            # Determine create vs update
            row_id = row.get("id", "").strip()
            if row_id:
                try:
                    existing = self._asset_repo.get(db, UUID(row_id))
                except (ValueError, Exception):
                    existing = None
                if existing:
                    item.action = ImportAction.UPDATE
                    item.existing_asset_id = str(existing.id)
                    result.will_update += 1
                else:
                    item.message = f"Asset with ID '{row_id}' not found (stale ID)"
                    result.errors += 1
                    result.items.append(item)
                    continue
            else:
                platform_val = row.get("platform", "").strip()
                location_val = row.get("location", "").strip()
                existing = self._asset_repo.get_by_identity(
                    db, name=name, asset_type_id=type_id,
                    platform=platform_val, location=location_val,
                )
                if existing:
                    item.action = ImportAction.UPDATE
                    item.existing_asset_id = str(existing.id)
                    result.will_update += 1
                else:
                    item.action = ImportAction.CREATE
                    result.will_create += 1

            result.items.append(item)

        return result

    # ------------------------------------------------------------------
    # Import - Execute
    # ------------------------------------------------------------------

    def execute_import(
        self, db: Session, *, file_bytes: bytes, filename: str, current_user_id: str
    ) -> ImportResult:
        """Parse, validate, and execute the import in a single transaction."""
        result = ImportResult()

        rows = self._parse_file(file_bytes, filename)
        if not rows:
            result.error_messages.append("File is empty or has no data rows.")
            return result

        if len(rows) > BULK_IMPORT_MAX_ROWS:
            result.error_messages.append(
                f"File contains {len(rows)} rows, maximum is {BULK_IMPORT_MAX_ROWS}. Split into multiple files."
            )
            return result

        type_map = self._resolve_asset_types(db)
        dup_first_rows = self._detect_duplicates(rows)
        cycles = self._detect_parent_cycles(rows)
        has_cycles = len(cycles) > 0

        # Track created assets for parent wiring
        created_assets: Dict[str, UUID] = {}  # name -> asset id

        for idx, row in enumerate(rows):
            row_num = idx + 1
            name = row.get("name", "").strip()
            asset_type_raw = row.get("asset_type", "").strip()
            item = ImportResultItem(row=row_num, name=name, asset_type=asset_type_raw, action=ImportAction.ERROR)

            if not name or not asset_type_raw:
                item.message = f"Missing required field: {'name' if not name else 'asset_type'}"
                result.errors += 1
                result.items.append(item)
                continue

            type_id = type_map.get(asset_type_raw.lower())
            if not type_id:
                item.message = f"Unknown asset type: '{asset_type_raw}'"
                result.errors += 1
                result.items.append(item)
                continue

            # Duplicate detection
            dup_key = "|".join([
                name.lower(), asset_type_raw.lower(),
                row.get("platform", "").strip().lower(),
                row.get("location", "").strip().lower(),
            ])
            if dup_key in dup_first_rows and (idx + 1) != dup_first_rows[dup_key]:
                item.message = f"Duplicate of row {dup_first_rows[dup_key]} in this file"
                result.errors += 1
                result.items.append(item)
                continue

            # Validate status
            status_raw = row.get("status", "").strip().lower()
            if status_raw and status_raw not in [s.value for s in AssetStatus]:
                item.message = f"Invalid status: '{status_raw}'"
                result.errors += 1
                result.items.append(item)
                continue

            # Parse properties
            try:
                properties = _parse_properties(row.get("properties", ""))
            except ValueError as e:
                item.message = str(e)
                result.errors += 1
                result.items.append(item)
                continue

            tags = _parse_tags(row.get("tags", ""))
            description = row.get("description", "").strip() or None
            platform_val = row.get("platform", "").strip() or None
            location_val = row.get("location", "").strip() or None
            domain_id_val = row.get("domain_id", "").strip() or None
            status_val = AssetStatus(status_raw) if status_raw else AssetStatus.ACTIVE

            # Determine create vs update
            row_id = row.get("id", "").strip()
            try:
                if row_id:
                    try:
                        existing = self._asset_repo.get(db, UUID(row_id))
                    except (ValueError, Exception):
                        existing = None
                    if not existing:
                        item.message = f"Asset with ID '{row_id}' not found"
                        result.errors += 1
                        result.items.append(item)
                        continue

                    update_data = AssetUpdate(
                        name=name,
                        description=description,
                        asset_type_id=type_id,
                        platform=platform_val,
                        location=location_val,
                        domain_id=domain_id_val,
                        properties=properties,
                        tags=tags,
                        status=status_val,
                    )
                    updated = self._asset_repo.update(
                        db=db, db_obj=existing, obj_in=update_data.model_dump(exclude_unset=True)
                    )
                    db.flush()
                    item.action = ImportAction.UPDATE
                    item.asset_id = str(updated.id)
                    result.updated += 1
                    created_assets[name.lower()] = updated.id
                else:
                    existing = self._asset_repo.get_by_identity(
                        db, name=name, asset_type_id=type_id,
                        platform=platform_val or "", location=location_val or "",
                    )
                    if existing:
                        update_data = AssetUpdate(
                            description=description,
                            domain_id=domain_id_val,
                            properties=properties,
                            tags=tags,
                            status=status_val,
                        )
                        updated = self._asset_repo.update(
                            db=db, db_obj=existing, obj_in=update_data.model_dump(exclude_unset=True)
                        )
                        db.flush()
                        item.action = ImportAction.UPDATE
                        item.asset_id = str(updated.id)
                        result.updated += 1
                        created_assets[name.lower()] = updated.id
                    else:
                        create_data = {
                            "name": name,
                            "description": description,
                            "asset_type_id": type_id,
                            "platform": platform_val,
                            "location": location_val,
                            "domain_id": domain_id_val,
                            "properties": properties,
                            "tags": tags,
                            "status": status_val.value,
                            "created_by": current_user_id,
                        }
                        db_asset = AssetDb(**create_data)
                        db.add(db_asset)
                        db.flush()
                        db.refresh(db_asset)
                        item.action = ImportAction.CREATE
                        item.asset_id = str(db_asset.id)
                        result.created += 1
                        created_assets[name.lower()] = db_asset.id

            except Exception as e:
                item.message = f"Error: {str(e)[:200]}"
                result.errors += 1
                result.items.append(item)
                continue

            result.items.append(item)

        # Wire parent relationships (skip if cycles detected)
        if not has_cycles:
            self._wire_parent_relationships(db, rows, type_map, created_assets, result)

        return result

    def _wire_parent_relationships(
        self,
        db: Session,
        rows: List[Dict[str, str]],
        type_map: Dict[str, UUID],
        created_assets: Dict[str, UUID],
        result: ImportResult,
    ) -> None:
        """Create parent-child relationships from parent_asset column."""
        for idx, row in enumerate(rows):
            parent_ref = row.get("parent_asset", "").strip()
            rel_type = row.get("parent_relationship_type", "").strip()
            name = row.get("name", "").strip()

            if not parent_ref or not name:
                continue
            if not rel_type:
                rel_type = "contains"

            child_id = created_assets.get(name.lower())
            if not child_id:
                continue

            # Resolve parent: first check created assets, then DB by name
            parent_id = created_assets.get(parent_ref.lower())
            if not parent_id:
                parent_assets = db.query(AssetDb).filter(
                    AssetDb.name.ilike(parent_ref)
                ).first()
                if parent_assets:
                    parent_id = parent_assets.id

            if not parent_id:
                logger.warning(f"Row {idx + 1}: Parent asset '{parent_ref}' not found, skipping relationship.")
                continue

            # Check depth limit
            depth = self._get_ancestor_depth(db, parent_id)
            if depth >= MAX_RELATIONSHIP_DEPTH:
                logger.warning(f"Row {idx + 1}: Max relationship depth ({MAX_RELATIONSHIP_DEPTH}) exceeded for parent '{parent_ref}'.")
                continue

            existing_rel = self._rel_repo.find_existing(
                db,
                source_asset_id=parent_id,
                target_asset_id=child_id,
                relationship_type=rel_type,
            )
            if not existing_rel:
                db_rel = AssetRelationshipDb(
                    source_asset_id=parent_id,
                    target_asset_id=child_id,
                    relationship_type=rel_type,
                    created_by="bulk-import",
                )
                db.add(db_rel)
                db.flush()

    def _get_ancestor_depth(self, db: Session, asset_id: UUID, max_depth: int = MAX_RELATIONSHIP_DEPTH) -> int:
        """Count how many ancestors an asset has (to prevent excessively deep hierarchies)."""
        depth = 0
        current_id = asset_id
        visited = set()
        while depth < max_depth:
            if current_id in visited:
                break
            visited.add(current_id)
            parent_rel = db.query(AssetRelationshipDb).filter(
                AssetRelationshipDb.target_asset_id == current_id
            ).first()
            if not parent_rel:
                break
            current_id = parent_rel.source_asset_id
            depth += 1
        return depth
